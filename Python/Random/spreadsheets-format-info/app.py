import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from thefuzz import fuzz
from thefuzz import process
import re

# Functie om e-mailadres om te zetten in een herkenbare naam
def extract_name_from_email(email):
    """
    Zet een e-mailadres om in een naam.
    Bijvoorbeeld: 'kmondria@example.com' wordt 'Klaas Mondria'
    """
    if not email or pd.isna(email):
        return None

    # Splits het gedeelte vóór het "@"
    local_part = email.split("@")[0]

    # Split op basis van punten (bijvoorbeeld "d.wolfs")
    if "." in local_part:
        name_parts = [part.capitalize() for part in local_part.split(".")]
        return " ".join(name_parts)

    # Anders (bijvoorbeeld letters zonder punt, zoals dwolfs)
    name_parts = re.findall('[a-zA-Z]+', local_part)  # Vind woorden
    if len(name_parts) == 2:  # Voorbeeld: ['d', 'wolfs']
        return f"{name_parts[0].capitalize()} {name_parts[1].capitalize()}"
    return " ".join(name_parts).capitalize()

# 1️⃣ Bestanden inlezen
def load_files(overzicht_file, antwoorden_file):
    """
    Lees bestanden in: overzicht en antwoorden.
    """
    overzicht_df = pd.read_excel(overzicht_file, engine="odf")
    antwoorden_df = pd.read_excel(antwoorden_file)
    antwoorden_df.columns = antwoorden_df.columns.str.strip()  # Strip spaties van kolomnamen
    return overzicht_df, antwoorden_df

# 2️⃣ Data combineren en verwerken
def process_data(overzicht_df, antwoorden_df, mapping):
    """
    Verwerk de data en gebruik fuzzy matching om de namen correct te koppelen.
    """
    # Genereer namen uit e-mailadressen
    antwoorden_df["Matched Naam"] = antwoorden_df["E-mailadres"].apply(extract_name_from_email)

    print("\n🔍 Namen in Overzicht.ods:")
    print(overzicht_df["Naam"].tolist())
    print("\n🔍 Namen gegenereerd uit Antwoorden.xlsx:")
    print(antwoorden_df["Matched Naam"].tolist())

    # Alle namen uit de overzichts- en antwoordenbestanden
    overzicht_namen = overzicht_df["Naam"].tolist()
    antwoorden_namen = antwoorden_df["Matched Naam"].tolist()

    # Gebruik fuzzy matching om overeenkomsten te vinden
    matched_namen = {}
    for overzicht_naam in overzicht_namen:
        # Zoek de beste match met een minimale score van 80
        beste_match = process.extractOne(overzicht_naam, antwoorden_namen, scorer=fuzz.token_sort_ratio)
        if beste_match and beste_match[1] >= 80:
            matched_namen[overzicht_naam] = beste_match[0]
        else:
            # Als er geen goede volledige match is, probeer de achternaam
            overzicht_achternaam = overzicht_naam.split()[-1]  # Neem de laatste naam
            beste_achternaam_match = process.extractOne(overzicht_achternaam, antwoorden_namen, scorer=fuzz.partial_ratio)
            if beste_achternaam_match and beste_achternaam_match[1] >= 80:
                matched_namen[overzicht_naam] = beste_achternaam_match[0]
            else:
                matched_namen[overzicht_naam] = None

    # Toon de fuzzy matches in de output
    print("\n🔗 Fuzzy Matches:")
    for overzicht_naam, matched_naam in matched_namen.items():
        print(f"- {overzicht_naam} → {matched_naam}")

    # Voeg fuzzy matches toe aan overzicht_df
    overzicht_df["Matched Naam"] = overzicht_df["Naam"].map(matched_namen)

    # Voeg een boolean-kolom toe om te meten wie de enquête heeft ingevuld
    overzicht_df["Heeft Enquête Ingevuld"] = overzicht_df["Matched Naam"].notna()

    # Correct ingevulde fuzzy matches ook toevoegen bij antwoorden_df
    antwoorden_df["Matched Naam"] = antwoorden_df["Matched Naam"].fillna("Geen informatie")

    # Merge de dataframes
    merged_df = pd.merge(overzicht_df, antwoorden_df, on="Matched Naam", how="left")

    # Vul de data in met behulp van het mapping-schema
    for enquête_vraag, overzicht_kolom in mapping.items():
        if enquête_vraag in antwoorden_df.columns and overzicht_kolom in overzicht_df.columns:
            print(f"Mapping: '{enquête_vraag}' → '{overzicht_kolom}'")
            # Vul data alleen voor personen waar de enquête is ingevuld
            mapping_filter = (
                (merged_df["Heeft Enquête Ingevuld"] == True) &  # Enquête wel ingevuld
                (merged_df[enquête_vraag].notna())              # Data is aanwezig
            )
            merged_df.loc[mapping_filter, overzicht_kolom] = merged_df.loc[mapping_filter, enquête_vraag]
        else:
            print(f"⚠️ Mapping-probleem: '{enquête_vraag}' of '{overzicht_kolom}' ontbreekt in de datasets!")

    # Vul ontbrekende waarden met lege waarde of NaT
    for col in merged_df.columns:
        if pd.api.types.is_datetime64_any_dtype(merged_df[col]):  # Datumtypes
            merged_df[col] = merged_df[col].fillna(pd.NaT)
        else:  # Andere types
            merged_df[col] = merged_df[col].fillna("Geen informatie")

    return merged_df

# 3️⃣ Kleuren toepassen
def apply_colors(merged_df, output_file):
    """
    Voeg kleuren toe: markeer rijen met ingevulde enquête.
    """
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Schrijf DataFrame naar Excel
        merged_df.to_excel(writer, index=False, sheet_name="Overzicht_Gevuld")
        workbook = writer.book
        sheet = writer.sheets["Overzicht_Gevuld"]

        # Donkergroene kleur
        green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")

        # Pas kleur toe op alle rijen met 'Heeft Enquête Ingevuld' = True
        for row in range(2, len(merged_df) + 2):  # Excel-rijen starten bij index 1
            if merged_df.iloc[row - 2]["Heeft Enquête Ingevuld"]:  # True
                for col in range(1, len(merged_df.columns) + 1):
                    sheet.cell(row=row, column=col).fill = green_fill

# Hoofdprogramma
if __name__ == "__main__":
    overzicht_file = "Overzicht.ods"
    antwoorden_file = "Antwoorden.xlsx"
    output_file = "Overzicht_Gevuld.xlsx"

    # Mapping van vragen naar kolommen in Overzicht
    mapping = {
        "Heb je een laptoptas?": "Laptoptas",
        "Heb je een laptopstandaard?": "Laptop standaard",
        "Welke beeldschermen heb je op werk?": "Beeldscherm werk",
        "Welke beeldschermen heb je thuis?": "Beeldscherm Thuis",
        "Heb je een toetsenbord?": "Toetsenbord - bedraad",
        "Dongel? (Een Dongle met USB-C, Ethernet en HDMI)": "Dongel",
        "Docking Station?": "Docking station",
        "Muis?": "Muis",
        "Koptelefoon?": "Koptelefoon",
        "Webcam?": "webcam",  # Webcam gematcht aan kolom
        "Mobiel?": "Mobiel",
    }

    print("✅ Bestanden inlezen...")
    try:
        overzicht_df, antwoorden_df = load_files(overzicht_file, antwoorden_file)
    except Exception as e:
        print(f"❌ Fout bij het inlezen van bestanden: {e}")
        exit()

    print("✅ Data verwerken...")
    try:
        merged_df = process_data(overzicht_df, antwoorden_df, mapping)
    except Exception as e:
        print(f"❌ Fout in gegevensverwerking: {e}")
        exit()

    print("✅ Kleuren toepassen en bestand opslaan...")
    try:
        apply_colors(merged_df, output_file)
        print(f"✅ Verwerkt bestand succesvol opgeslagen als {output_file}")
    except Exception as e:
        print(f"❌ Fout bij opslaan: {e}")

    # Debugging: Controleer de gematchte en verwerkte data
    print("\n✅ Gematchte data (voorbeeld):")
    print(merged_df[["Naam", "Matched Naam", "Heeft Enquête Ingevuld"] + list(mapping.values())].head(10))